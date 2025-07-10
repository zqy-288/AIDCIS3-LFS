"""
独立的检测报告生成器模块（无GUI依赖）
专门用于测试环境，移除了PySide6依赖
支持多种形式的检测报告生成：静态PDF、动态Web、原始数据导出
"""

import os
import sys
import json
import csv
import tempfile
import threading
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

# 数据处理
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import numpy as np

# 报告生成（需要安装）
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm, mm
    from reportlab.lib.colors import Color, black, red, green, blue
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.platypus import Image as RLImage, PageBreak, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ reportlab未安装，PDF报告功能不可用")

# Excel导出（需要安装）
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl未安装，Excel导出功能不可用")

# 设置中文字体支持
def setup_chinese_fonts():
    """设置matplotlib中文字体支持"""
    import matplotlib.font_manager as fm
    
    # 查找系统中的中文字体
    chinese_fonts = []
    for font in fm.fontManager.ttflist:
        if 'PingFang' in font.name or 'Hiragino' in font.name or 'STHeiti' in font.name or 'Arial Unicode MS' in font.name:
            chinese_fonts.append(font.name)
            print(f"添加字体: {font.fname}")
    
    if chinese_fonts:
        plt.rcParams['font.sans-serif'] = chinese_fonts + ['DejaVu Sans', 'Arial']
        plt.rcParams['axes.unicode_minus'] = False
        print(f"使用中文字体: {chinese_fonts[0]}")
    else:
        print("⚠️ 未找到中文字体，使用默认字体")

# 初始化字体
setup_chinese_fonts()


class MockSignal:
    """模拟Qt信号类"""
    def __init__(self):
        self._handlers = []
    
    def connect(self, handler):
        self._handlers.append(handler)
    
    def emit(self, *args, **kwargs):
        for handler in self._handlers:
            try:
                handler(*args, **kwargs)
            except Exception as e:
                pass  # 静默处理信号错误


class ReportData:
    """报告数据容器类"""
    
    def __init__(self):
        # 基本信息
        self.report_id = ""
        self.workpiece_model = ""
        self.workpiece_serial = ""
        self.operator = ""
        self.start_time = None
        self.end_time = None
        self.equipment_id = ""
        
        # 统计数据
        self.total_holes = 0
        self.checked_holes = 0
        self.qualified_holes = 0
        self.unqualified_holes = 0
        self.qualification_rate = 0.0
        self.overall_result = "合格"
        
        # 详细数据
        self.non_conformities = []  # 不合格项列表
        self.all_data = []          # 全部检测数据
        self.charts = {}            # 图表文件路径
        self.images = {}            # 图像文件路径
        
    def add_non_conformity(self, hole_id: str, problem_type: str, 
                          measurement_result: str, evidence: Dict, 
                          review_record: Dict):
        """添加不合格项"""
        nc = {
            'hole_id': hole_id,
            'problem_type': problem_type,
            'measurement_result': measurement_result,
            'evidence': evidence,
            'review_record': review_record,
            'timestamp': datetime.now()
        }
        self.non_conformities.append(nc)
        
    def calculate_summary(self):
        """计算摘要统计信息"""
        if self.all_data:
            self.checked_holes = len(self.all_data)
            qualified_count = sum(1 for item in self.all_data if item.get('qualified', True))
            self.qualified_holes = qualified_count
            self.unqualified_holes = self.checked_holes - qualified_count
            
            if self.checked_holes > 0:
                self.qualification_rate = round((qualified_count / self.checked_holes) * 100, 2)
            else:
                self.qualification_rate = 0.0
                
            # 判断总体结果
            self.overall_result = "合格" if self.unqualified_holes == 0 else "不合格"


class ReportGenerator:
    """报告生成器类"""
    
    def __init__(self, db_manager=None):
        self.db_manager = db_manager
        self.temp_dir = Path(tempfile.mkdtemp())
        self.output_dir = Path("reports")
        self.output_dir.mkdir(exist_ok=True)
        
        self.company_name = "数字化检测系统"
        
        # 中文字体设置
        self.chinese_fonts = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
        
    def _prepare_report_data(self, hole_data: Dict, workpiece_info: Dict) -> ReportData:
        """准备报告数据"""
        report_data = ReportData()
        
        # 设置基本信息
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        model = workpiece_info.get('model', 'Unknown')
        report_data.report_id = f"REP-{model}-{timestamp}"
        
        report_data.workpiece_model = model
        report_data.workpiece_serial = workpiece_info.get('serial', 'Unknown')
        report_data.operator = workpiece_info.get('operator', 'Unknown')
        report_data.start_time = workpiece_info.get('start_time', datetime.now())
        report_data.end_time = workpiece_info.get('end_time', datetime.now())
        report_data.equipment_id = workpiece_info.get('equipment_id', 'AIDCIS-001')
        
        # 设置统计数据
        report_data.total_holes = hole_data.get('total_holes', 0)
        report_data.checked_holes = hole_data.get('checked_holes', report_data.total_holes)
        report_data.qualified_holes = hole_data.get('qualified_holes', report_data.total_holes - 5)
        report_data.unqualified_holes = hole_data.get('unqualified_holes', 5)
        
        if report_data.checked_holes > 0:
            report_data.qualification_rate = round(
                (report_data.qualified_holes / report_data.checked_holes) * 100, 2
            )
        else:
            report_data.qualification_rate = 0.0
            
        report_data.overall_result = "合格" if report_data.unqualified_holes == 0 else "不合格"
        
        # 生成模拟数据
        self._generate_mock_data(report_data)
        
        return report_data
        
    def _generate_mock_data(self, report_data: ReportData):
        """生成模拟检测数据"""
        # 生成全部检测数据
        for i in range(report_data.total_holes):
            hole_id = f"H{i+1:03d}"
            is_qualified = i >= report_data.unqualified_holes  # 前几个设为不合格
            
            hole_info = {
                'hole_id': hole_id,
                'min_diameter': round(17.55 + np.random.normal(0, 0.005), 3),
                'max_diameter': round(17.65 + np.random.normal(0, 0.005), 3),
                'avg_diameter': round(17.60 + np.random.normal(0, 0.005), 3),
                'qualified': is_qualified,
                'surface_defects': 'None' if is_qualified else 'Diameter deviation',
                'position_x': (i % 10) * 15,
                'position_y': (i // 10) * 15
            }
            
            report_data.all_data.append(hole_info)
            
            # 为不合格孔位添加不合格项
            if not is_qualified:
                report_data.add_non_conformity(
                    hole_id=hole_id,
                    problem_type="孔径超差",
                    measurement_result=f"{hole_info['avg_diameter']}mm",
                    evidence={'chart': f'envelope_{hole_id}.png'},
                    review_record={'reviewer': '质检员', 'time': datetime.now()}
                )
                
    def export_raw_data_csv(self, hole_data: Dict, workpiece_info: Dict) -> str:
        """导出原始数据为CSV格式"""
        # 准备报告数据
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"检测数据_{report_data.workpiece_model}_{timestamp}.csv"
        csv_path = self.output_dir / filename
        
        # 写入CSV文件
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as csvfile:
            fieldnames = [
                '孔位ID', '最小直径(mm)', '最大直径(mm)', '平均直径(mm)',
                '检测状态', '表面缺陷', '位置X(mm)', '位置Y(mm)'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for item in report_data.all_data:
                writer.writerow({
                    '孔位ID': item['hole_id'],
                    '最小直径(mm)': item['min_diameter'],
                    '最大直径(mm)': item['max_diameter'],
                    '平均直径(mm)': item['avg_diameter'],
                    '检测状态': '合格' if item['qualified'] else '不合格',
                    '表面缺陷': item['surface_defects'],
                    '位置X(mm)': item['position_x'],
                    '位置Y(mm)': item['position_y']
                })
                
        return str(csv_path)
        
    def export_raw_data_excel(self, hole_data: Dict, workpiece_info: Dict) -> str:
        """导出原始数据为Excel格式"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库不可用，无法生成Excel报告")
            
        # 准备报告数据
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"检测数据_{report_data.workpiece_model}_{timestamp}.xlsx"
        excel_path = self.output_dir / filename
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 创建工作表
        ws_summary = wb.active
        ws_summary.title = "检测摘要"
        
        # 写入摘要信息
        summary_data = [
            ['报告ID', report_data.report_id],
            ['工件型号', report_data.workpiece_model],
            ['工件序列号', report_data.workpiece_serial],
            ['操作员', report_data.operator],
            ['检测设备', report_data.equipment_id],
            ['总孔数', report_data.total_holes],
            ['已检孔数', report_data.checked_holes],
            ['合格孔数', report_data.qualified_holes],
            ['不合格孔数', report_data.unqualified_holes],
            ['合格率(%)', report_data.qualification_rate],
            ['总体结果', report_data.overall_result]
        ]
        
        for row_idx, (key, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=key)
            ws_summary.cell(row=row_idx, column=2, value=value)
            
        # 创建详细数据工作表
        ws_data = wb.create_sheet("详细数据")
        
        # 写入表头
        headers = ['孔位ID', '最小直径(mm)', '最大直径(mm)', '平均直径(mm)', 
                  '检测状态', '表面缺陷', '位置X(mm)', '位置Y(mm)']
        
        for col_idx, header in enumerate(headers, 1):
            ws_data.cell(row=1, column=col_idx, value=header)
            
        # 写入数据
        for row_idx, item in enumerate(report_data.all_data, 2):
            ws_data.cell(row=row_idx, column=1, value=item['hole_id'])
            ws_data.cell(row=row_idx, column=2, value=item['min_diameter'])
            ws_data.cell(row=row_idx, column=3, value=item['max_diameter'])
            ws_data.cell(row=row_idx, column=4, value=item['avg_diameter'])
            ws_data.cell(row=row_idx, column=5, value='合格' if item['qualified'] else '不合格')
            ws_data.cell(row=row_idx, column=6, value=item['surface_defects'])
            ws_data.cell(row=row_idx, column=7, value=item['position_x'])
            ws_data.cell(row=row_idx, column=8, value=item['position_y'])
            
        # 保存文件
        wb.save(excel_path)
        
        return str(excel_path)
        
    def generate_web_report_data(self, hole_data: Dict, workpiece_info: Dict) -> Dict:
        """生成Web报告数据"""
        # 准备报告数据
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 构造Web报告数据结构
        web_data = {
            'header': {
                'report_id': report_data.report_id,
                'workpiece_model': report_data.workpiece_model,
                'workpiece_serial': report_data.workpiece_serial,
                'operator': report_data.operator,
                'equipment_id': report_data.equipment_id,
                'start_time': report_data.start_time.isoformat() if report_data.start_time else None,
                'end_time': report_data.end_time.isoformat() if report_data.end_time else None,
                'generated_at': datetime.now().isoformat()
            },
            'summary': {
                'total_holes': report_data.total_holes,
                'checked_holes': report_data.checked_holes,
                'qualified_holes': report_data.qualified_holes,
                'unqualified_holes': report_data.unqualified_holes,
                'qualification_rate': report_data.qualification_rate,
                'overall_result': report_data.overall_result
            },
            'non_conformities': [
                {
                    'hole_id': nc['hole_id'],
                    'problem_type': nc['problem_type'],
                    'measurement_result': nc['measurement_result'],
                    'evidence': nc['evidence'],
                    'review_record': nc['review_record']
                }
                for nc in report_data.non_conformities
            ],
            'charts': report_data.charts,
            'images': report_data.images,
            'full_data': report_data.all_data
        }
        
        return web_data
        
    def generate_envelope_chart(self, measurement_data: List[Dict], 
                              target_diameter: float, upper_tolerance: float, 
                              lower_tolerance: float) -> str:
        """生成包络图"""
        if not measurement_data:
            # 生成模拟数据
            measurement_data = [
                {'depth': i * 0.1, 'diameter': target_diameter + np.random.normal(0, 0.005)}
                for i in range(1000)
            ]
            
        # 提取数据
        depths = [d.get('depth', i * 0.1) for i, d in enumerate(measurement_data)]
        diameters = [d.get('diameter', target_diameter) for d in measurement_data]
        
        # 创建图形
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # 绘制测量数据
        ax.plot(depths, diameters, 'b-', linewidth=1.5, label='实测直径')
        
        # 计算公差带
        upper_limit = target_diameter + upper_tolerance
        lower_limit = target_diameter - lower_tolerance
        
        # 绘制公差线
        ax.axhline(y=upper_limit, color='red', linestyle='--', linewidth=2, 
                  label=f'上限 {upper_limit:.3f}mm')
        ax.axhline(y=lower_limit, color='red', linestyle='--', linewidth=2, 
                  label=f'下限 {lower_limit:.3f}mm')
        ax.axhline(y=target_diameter, color='green', linestyle='-', linewidth=2, 
                  label=f'目标直径 {target_diameter:.3f}mm')
        
        # 填充合格区域
        ax.fill_between(depths, upper_limit, lower_limit, alpha=0.2, color='green')
        
        # 设置图形属性
        ax.set_xlabel('探头深度 (mm)', fontsize=12)
        ax.set_ylabel('孔径直径 (mm)', fontsize=12)
        ax.set_title('孔径包络图', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.legend()
        
        # 保存图片
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        chart_filename = f"envelope_chart_{timestamp}.png"
        chart_path = self.temp_dir / chart_filename
        
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return str(chart_path)


class ReportGenerationThread:
    """模拟报告生成线程（无Qt依赖）"""
    
    def __init__(self, generator: ReportGenerator, report_type: str, 
                 hole_data: Dict, workpiece_info: Dict):
        self.generator = generator
        self.report_type = report_type
        self.hole_data = hole_data
        self.workpiece_info = workpiece_info
        
        # 模拟信号
        self.progress_updated = MockSignal()
        self.status_updated = MockSignal()
        self.generation_completed = MockSignal()
        self.generation_failed = MockSignal()
        
        self._thread = None
        self._running = False
        
    def start(self):
        """启动线程"""
        self._thread = threading.Thread(target=self.run, daemon=True)
        self._running = True
        self._thread.start()
        
    def run(self):
        """运行报告生成"""
        try:
            self.status_updated.emit("开始生成报告...")
            self.progress_updated.emit(10)
            
            if self.report_type == "CSV":
                self.status_updated.emit("正在生成CSV报告...")
                file_path = self.generator.export_raw_data_csv(
                    self.hole_data, self.workpiece_info
                )
                self.progress_updated.emit(50)
                
            elif self.report_type == "EXCEL":
                self.status_updated.emit("正在生成Excel报告...")
                file_path = self.generator.export_raw_data_excel(
                    self.hole_data, self.workpiece_info
                )
                self.progress_updated.emit(50)
                
            elif self.report_type == "WEB":
                self.status_updated.emit("正在生成Web报告...")
                web_data = self.generator.generate_web_report_data(
                    self.hole_data, self.workpiece_info
                )
                
                # 保存Web数据为JSON文件
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                model = self.workpiece_info.get('model', 'Unknown')
                filename = f"web_report_{model}_{timestamp}.json"
                file_path = self.generator.output_dir / filename
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(web_data, f, ensure_ascii=False, indent=2, default=str)
                    
                file_path = str(file_path)
                self.progress_updated.emit(50)
                
            else:
                raise ValueError(f"不支持的报告类型: {self.report_type}")
                
            self.progress_updated.emit(100)
            self.status_updated.emit("报告生成完成")
            self.generation_completed.emit(self.report_type, file_path)
            
        except Exception as e:
            self.generation_failed.emit(self.report_type, str(e))
        finally:
            self._running = False
            
    def wait(self, timeout: int = 5000):
        """等待线程完成"""
        if self._thread:
            timeout_seconds = timeout / 1000.0
            self._thread.join(timeout=timeout_seconds)
            return not self._thread.is_alive()
        return True


if __name__ == "__main__":
    """测试报告生成器"""
    print("测试独立报告生成器...")
    
    # 创建生成器
    generator = ReportGenerator()
    
    # 测试数据
    hole_data = {
        'total_holes': 50,
        'checked_holes': 48,
        'qualified_holes': 45,
        'unqualified_holes': 3
    }
    
    workpiece_info = {
        'model': 'CP1400',
        'serial': 'SN-TEST-001',
        'operator': '测试用户',
        'start_time': datetime.now(),
        'end_time': datetime.now()
    }
    
    # 测试CSV导出
    print("测试CSV导出...")
    csv_path = generator.export_raw_data_csv(hole_data, workpiece_info)
    print(f"CSV文件已生成: {csv_path}")
    
    # 测试Web数据生成
    print("测试Web数据生成...")
    web_data = generator.generate_web_report_data(hole_data, workpiece_info)
    print(f"Web数据已生成: {len(web_data)} 个字段")
    
    # 测试包络图生成
    print("测试包络图生成...")
    measurement_data = [
        {'depth': i * 0.5, 'diameter': 17.6 + np.random.normal(0, 0.005)}
        for i in range(200)
    ]
    chart_path = generator.generate_envelope_chart(measurement_data, 17.6, 0.05, 0.07)
    print(f"包络图已生成: {chart_path}")
    
    print("✅ 独立报告生成器测试完成")