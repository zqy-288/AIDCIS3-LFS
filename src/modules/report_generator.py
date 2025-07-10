"""
检测报告生成器模块
支持多种形式的检测报告生成：静态PDF、动态Web、原始数据导出
按照甲方要求的报告框架结构实现
"""

import os
import sys
import json
import csv
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

# Qt相关
from PySide6.QtWidgets import QWidget, QMessageBox, QProgressDialog, QApplication
from PySide6.QtCore import QThread, Signal, QTimer

# 数据处理
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
import numpy as np

# 报告生成（需要安装）
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm, mm
    from reportlab.lib.colors import Color, black, red, green, blue
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.platypus import Image as RLImage, PageBreak, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ reportlab未安装，PDF报告功能不可用")

# Excel导出（需要安装）
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl未安装，Excel导出功能不可用")


class ReportData:
    """报告数据结构"""
    
    def __init__(self):
        # 报告头信息
        self.report_id = ""
        self.workpiece_model = ""
        self.workpiece_serial = ""
        self.detection_equipment = ""
        self.start_time = None
        self.end_time = None
        self.operator = ""
        self.overall_result = "合格"  # 合格/不合格
        
        # 摘要信息
        self.total_holes = 0
        self.checked_holes = 0
        self.qualified_holes = 0
        self.unqualified_holes = 0
        self.qualification_rate = 0.0
        
        # 不合格项列表
        self.non_conformities = []
        
        # 所有检测数据
        self.all_data = []
        
        # 图表和图像
        self.charts = {}  # 图表文件路径
        self.images = {}  # 图像文件路径
        
    def add_non_conformity(self, hole_id: str, problem_type: str, 
                          measurement_result: str, evidence: Dict, 
                          review_record: Dict = None):
        """添加不合格项"""
        nc = {
            'hole_id': hole_id,
            'problem_type': problem_type,
            'measurement_result': measurement_result,
            'evidence': evidence,
            'review_record': review_record or {}
        }
        self.non_conformities.append(nc)
        
    def calculate_summary(self):
        """计算摘要信息"""
        self.checked_holes = len(self.all_data)
        self.unqualified_holes = len(self.non_conformities)
        self.qualified_holes = self.checked_holes - self.unqualified_holes
        if self.checked_holes > 0:
            self.qualification_rate = (self.qualified_holes / self.checked_holes) * 100
        self.overall_result = "合格" if self.unqualified_holes == 0 else "不合格"


class ReportGenerator:
    """检测报告生成器主类"""
    
    def __init__(self, db_manager=None):
        self.db_manager = db_manager
        self.temp_dir = Path("temp_reports")
        self.temp_dir.mkdir(exist_ok=True)
        self.output_dir = Path("reports")
        self.output_dir.mkdir(exist_ok=True)
        
        # 报告配置
        self.company_name = "数字化检测系统"
        self.company_logo = None  # 公司Logo路径
        
    def generate_static_pdf_report(self, hole_data: Dict, workpiece_info: Dict) -> str:
        """
        生成静态PDF报告
        
        Args:
            hole_data: 孔位检测数据
            workpiece_info: 工件基本信息
            
        Returns:
            生成的PDF文件路径
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("需要安装reportlab库才能生成PDF报告")
            
        # 准备报告数据
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 生成报告文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"检测报告_{report_data.workpiece_model}_{timestamp}.pdf"
        report_path = self.output_dir / report_filename
        
        # 创建PDF文档
        doc = SimpleDocTemplate(
            str(report_path),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # 构建报告内容
        story = []
        styles = getSampleStyleSheet()
        
        # 添加自定义样式
        self._add_custom_styles(styles)
        
        # 1. 封面/报告头
        story.extend(self._build_cover_section(report_data, styles))
        
        # 2. 摘要信息
        story.extend(self._build_summary_section(report_data, styles))
        
        # 3. 不合格项详单
        if report_data.non_conformities:
            story.extend(self._build_non_conformities_section(report_data, styles))
        
        # 4. 全部检测数据列表（可选）
        story.extend(self._build_full_data_section(report_data, styles))
        
        # 5. 页脚与签核区
        story.extend(self._build_footer_section(report_data, styles))
        
        # 生成PDF
        doc.build(story)
        
        print(f"✅ PDF报告生成完成: {report_path}")
        return str(report_path)
        
    def generate_web_report_data(self, hole_data: Dict, workpiece_info: Dict) -> Dict:
        """
        生成Web报告数据（JSON格式）
        
        Args:
            hole_data: 孔位检测数据
            workpiece_info: 工件基本信息
            
        Returns:
            Web报告数据字典
        """
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 转换为Web友好的数据格式
        web_data = {
            'header': {
                'report_id': report_data.report_id,
                'workpiece_model': report_data.workpiece_model,
                'workpiece_serial': report_data.workpiece_serial,
                'detection_equipment': report_data.detection_equipment,
                'start_time': report_data.start_time.isoformat() if report_data.start_time else None,
                'end_time': report_data.end_time.isoformat() if report_data.end_time else None,
                'operator': report_data.operator,
                'overall_result': report_data.overall_result
            },
            'summary': {
                'total_holes': report_data.total_holes,
                'checked_holes': report_data.checked_holes,
                'qualified_holes': report_data.qualified_holes,
                'unqualified_holes': report_data.unqualified_holes,
                'qualification_rate': round(report_data.qualification_rate, 2)
            },
            'non_conformities': report_data.non_conformities,
            'charts': report_data.charts,
            'images': report_data.images,
            'full_data': report_data.all_data[:100]  # 限制Web显示数据量
        }
        
        return web_data
        
    def export_raw_data_excel(self, hole_data: Dict, workpiece_info: Dict) -> str:
        """
        导出原始数据为Excel格式
        
        Args:
            hole_data: 孔位检测数据
            workpiece_info: 工件基本信息
            
        Returns:
            生成的Excel文件路径
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("需要安装openpyxl库才能导出Excel文件")
            
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"检测数据_{report_data.workpiece_model}_{timestamp}.xlsx"
        excel_path = self.output_dir / excel_filename
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 1. 基本信息工作表
        ws_info = wb.active
        ws_info.title = "基本信息"
        self._write_excel_info_sheet(ws_info, report_data)
        
        # 2. 检测数据工作表
        ws_data = wb.create_sheet("检测数据")
        self._write_excel_data_sheet(ws_data, report_data)
        
        # 3. 统计分析工作表
        ws_stats = wb.create_sheet("统计分析")
        self._write_excel_stats_sheet(ws_stats, report_data)
        
        # 4. 不合格项工作表
        if report_data.non_conformities:
            ws_nc = wb.create_sheet("不合格项")
            self._write_excel_nc_sheet(ws_nc, report_data)
        
        # 保存文件
        wb.save(excel_path)
        
        print(f"✅ Excel数据导出完成: {excel_path}")
        return str(excel_path)
        
    def export_raw_data_csv(self, hole_data: Dict, workpiece_info: Dict) -> str:
        """
        导出原始数据为CSV格式
        
        Args:
            hole_data: 孔位检测数据
            workpiece_info: 工件基本信息
            
        Returns:
            生成的CSV文件路径
        """
        report_data = self._prepare_report_data(hole_data, workpiece_info)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"检测数据_{report_data.workpiece_model}_{timestamp}.csv"
        csv_path = self.output_dir / csv_filename
        
        # 写入CSV文件
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # 写入表头
            headers = [
                '孔位ID', '最小直径(mm)', '最大直径(mm)', '平均直径(mm)', 
                '椭圆度', '表面缺陷', '最终状态', '检测时间'
            ]
            writer.writerow(headers)
            
            # 写入数据
            for data in report_data.all_data:
                row = [
                    data.get('hole_id', ''),
                    data.get('min_diameter', ''),
                    data.get('max_diameter', ''),
                    data.get('avg_diameter', ''),
                    data.get('ellipticity', ''),
                    '有' if data.get('has_defect', False) else '无',
                    '合格' if data.get('qualified', True) else '不合格',
                    data.get('detection_time', '')
                ]
                writer.writerow(row)
        
        print(f"✅ CSV数据导出完成: {csv_path}")
        return str(csv_path)
        
    def generate_envelope_chart(self, measurement_data: List[Dict], 
                               target_diameter: float, 
                               upper_tolerance: float, 
                               lower_tolerance: float) -> str:
        """
        生成包络图（公差带包络图）
        
        Args:
            measurement_data: 测量数据列表
            target_diameter: 目标直径
            upper_tolerance: 上公差
            lower_tolerance: 下公差
            
        Returns:
            生成的图片文件路径
        """
        # 准备数据
        depths = [d['depth'] for d in measurement_data]
        diameters = [d['diameter'] for d in measurement_data]
        
        # 创建图形
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # 绘制测量数据
        ax.plot(depths, diameters, 'b-', linewidth=2, label='测量数据')
        
        # 绘制公差带
        upper_limit = target_diameter + upper_tolerance
        lower_limit = target_diameter - lower_tolerance
        
        ax.axhline(y=upper_limit, color='r', linestyle='--', linewidth=2, 
                  label=f'上限 {upper_limit:.3f}mm')
        ax.axhline(y=lower_limit, color='r', linestyle='--', linewidth=2, 
                  label=f'下限 {lower_limit:.3f}mm')
        ax.axhline(y=target_diameter, color='g', linestyle='-', linewidth=1, 
                  label=f'目标直径 {target_diameter:.3f}mm')
        
        # 填充公差带
        ax.fill_between(depths, upper_limit, lower_limit, alpha=0.2, color='green', 
                       label='合格区域')
        
        # 标记超差点
        for i, (depth, diameter) in enumerate(zip(depths, diameters)):
            if diameter > upper_limit or diameter < lower_limit:
                ax.scatter(depth, diameter, color='red', s=50, zorder=5)
                ax.annotate(f'{diameter:.3f}mm', (depth, diameter), 
                           xytext=(5, 5), textcoords='offset points',
                           bbox=dict(boxstyle='round,pad=0.3', facecolor='red', alpha=0.7),
                           color='white', fontweight='bold')
        
        # 设置标签和标题
        ax.set_xlabel('深度 (mm)', fontsize=12, fontweight='bold')
        ax.set_ylabel('直径 (mm)', fontsize=12, fontweight='bold')
        ax.set_title('孔径包络图 - 公差带分析', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.legend(fontsize=10)
        
        # 保存图片
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        chart_filename = f"envelope_chart_{timestamp}.png"
        chart_path = self.temp_dir / chart_filename
        
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return str(chart_path)
        
    def _prepare_report_data(self, hole_data: Dict, workpiece_info: Dict) -> ReportData:
        """准备报告数据"""
        report_data = ReportData()
        
        # 填充基本信息
        timestamp = datetime.now()
        report_data.report_id = f"REP-{workpiece_info.get('model', 'UNKNOWN')}-{timestamp.strftime('%Y%m%d-%H%M%S')}"
        report_data.workpiece_model = workpiece_info.get('model', 'CP1400')
        report_data.workpiece_serial = workpiece_info.get('serial', f"SN-{timestamp.strftime('%Y%m%d-%H%M')}")
        report_data.detection_equipment = "数字化检测工作站 V1.0"
        report_data.start_time = workpiece_info.get('start_time', timestamp)
        report_data.end_time = workpiece_info.get('end_time', timestamp)
        report_data.operator = workpiece_info.get('operator', '系统用户')
        
        # 填充检测数据（示例数据，实际应从数据库获取）
        report_data.total_holes = hole_data.get('total_holes', 1500)
        
        # 模拟全部检测数据
        for i in range(1, report_data.total_holes + 1):
            hole_id = f"C{i//50+1:02d}R{i%50+1:02d}"
            data = {
                'hole_id': hole_id,
                'min_diameter': 16.95 + np.random.normal(0, 0.02),
                'max_diameter': 17.05 + np.random.normal(0, 0.02),
                'avg_diameter': 17.00 + np.random.normal(0, 0.01),
                'ellipticity': abs(np.random.normal(0, 0.005)),
                'has_defect': np.random.random() < 0.01,  # 1%概率有缺陷
                'qualified': True,
                'detection_time': timestamp.strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # 随机生成一些不合格项
            if np.random.random() < 0.002:  # 0.2%概率不合格
                data['qualified'] = False
                if data['min_diameter'] < 16.95:
                    report_data.add_non_conformity(
                        hole_id=hole_id,
                        problem_type="孔径测量不合格",
                        measurement_result=f"最小内径 {data['min_diameter']:.3f}mm (标准: 17.00mm ~ 17.12mm)",
                        evidence={'envelope_chart': 'path_to_chart.png'},
                        review_record={
                            'reviewer': '张工',
                            'review_time': timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                            'review_result': f"{data['min_diameter']:.3f}mm",
                            'final_decision': '确认不合格'
                        }
                    )
            
            report_data.all_data.append(data)
        
        # 计算摘要
        report_data.calculate_summary()
        
        return report_data
        
    def _add_custom_styles(self, styles):
        """添加自定义样式"""
        # 报告标题样式
        styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=black,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # 章节标题样式
        styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=blue,
            spaceBefore=20,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        ))
        
        # 不合格项标题样式
        styles.add(ParagraphStyle(
            name='NonConformityTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=red,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        ))
        
    def _build_cover_section(self, report_data: ReportData, styles) -> List:
        """构建封面/报告头部分"""
        story = []
        
        # 报告标题
        title = f"{report_data.workpiece_model}型号管板数字化检测报告"
        story.append(Paragraph(title, styles['ReportTitle']))
        story.append(Spacer(1, 20))
        
        # 基本信息表格
        info_data = [
            ['报告ID:', report_data.report_id],
            ['工件信息:', ''],
            ['　产品型号:', report_data.workpiece_model],
            ['　工件序列号:', report_data.workpiece_serial],
            ['检测信息:', ''],
            ['　检测设备编号:', report_data.detection_equipment],
            ['　检测开始时间:', report_data.start_time.strftime('%Y-%m-%d %H:%M:%S') if report_data.start_time else ''],
            ['　检测结束时间:', report_data.end_time.strftime('%Y-%m-%d %H:%M:%S') if report_data.end_time else ''],
            ['　操作员:', report_data.operator],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, black),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # 最终结论
        result_color = green if report_data.overall_result == "合格" else red
        result_text = f"<font size=24 color={result_color.hexval()}><b>最终结论: {report_data.overall_result}</b></font>"
        story.append(Paragraph(result_text, styles['Normal']))
        story.append(PageBreak())
        
        return story
        
    def _build_summary_section(self, report_data: ReportData, styles) -> List:
        """构建摘要信息部分"""
        story = []
        
        story.append(Paragraph("摘要信息", styles['SectionTitle']))
        
        # 总体检测概览
        summary_data = [
            ['项目', '数值'],
            ['总孔数', str(report_data.total_holes)],
            ['已检孔数', str(report_data.checked_holes)],
            ['合格孔数', str(report_data.qualified_holes)],
            ['不合格孔数', str(report_data.unqualified_holes)],
            ['总体合格率', f"{report_data.qualification_rate:.2f}%"],
        ]
        
        summary_table = Table(summary_data, colWidths=[6*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BACKGROUND', (0, 0), (-1, 0), Color(0.8, 0.8, 0.8)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, black),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # 不合格项列表
        if report_data.non_conformities:
            story.append(Paragraph("不合格项概览:", styles['Normal']))
            for i, nc in enumerate(report_data.non_conformities, 1):
                nc_text = f"{i}. {nc['hole_id']}: {nc['problem_type']}"
                story.append(Paragraph(nc_text, styles['Normal']))
        else:
            story.append(Paragraph("✓ 所有检测项目均合格", styles['Normal']))
            
        story.append(Spacer(1, 20))
        
        return story
        
    def _build_non_conformities_section(self, report_data: ReportData, styles) -> List:
        """构建不合格项详单部分"""
        story = []
        
        story.append(Paragraph("不合格项详单", styles['SectionTitle']))
        
        for i, nc in enumerate(report_data.non_conformities, 1):
            # 不合格项标题
            nc_title = f"不合格项 {i}："
            story.append(Paragraph(nc_title, styles['NonConformityTitle']))
            
            # 详细信息
            nc_info = [
                ['孔位ID:', nc['hole_id']],
                ['问题类型:', nc['problem_type']],
                ['机检结果:', nc['measurement_result']],
            ]
            
            # 人工复检记录
            if nc.get('review_record'):
                review = nc['review_record']
                nc_info.extend([
                    ['人工复检记录:', ''],
                    ['　复检人:', review.get('reviewer', '')],
                    ['　复检时间:', review.get('review_time', '')],
                    ['　复检结果:', review.get('review_result', '')],
                    ['　最终判定:', review.get('final_decision', '')],
                ])
            
            nc_table = Table(nc_info, colWidths=[4*cm, 8*cm])
            nc_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, black),
            ]))
            
            story.append(nc_table)
            story.append(Spacer(1, 15))
            
        return story
        
    def _build_full_data_section(self, report_data: ReportData, styles) -> List:
        """构建全部检测数据列表部分"""
        story = []
        
        story.append(Paragraph("全部检测数据列表", styles['SectionTitle']))
        story.append(Paragraph("(仅显示前50项，完整数据请参考附件)", styles['Normal']))
        story.append(Spacer(1, 10))
        
        # 数据表格头
        data_headers = ['孔位ID', '最小直径(mm)', '最大直径(mm)', '平均直径(mm)', '椭圆度', '表面缺陷', '最终状态']
        data_table_data = [data_headers]
        
        # 添加数据行（限制显示数量）
        for data in report_data.all_data[:50]:
            row = [
                data['hole_id'],
                f"{data['min_diameter']:.3f}",
                f"{data['max_diameter']:.3f}",
                f"{data['avg_diameter']:.3f}",
                f"{data['ellipticity']:.3f}",
                '有' if data['has_defect'] else '无',
                '合格' if data['qualified'] else '不合格'
            ]
            data_table_data.append(row)
        
        data_table = Table(data_table_data, colWidths=[2*cm, 2*cm, 2*cm, 2*cm, 1.5*cm, 1.5*cm, 1.5*cm])
        data_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), Color(0.8, 0.8, 0.8)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, black),
        ]))
        
        story.append(data_table)
        story.append(Spacer(1, 20))
        
        return story
        
    def _build_footer_section(self, report_data: ReportData, styles) -> List:
        """构建页脚与签核区部分"""
        story = []
        
        # 报告生成信息
        generation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        footer_info = f"""
        <para align=center>
        报告生成时间: {generation_time}<br/>
        生成系统: {report_data.detection_equipment}<br/><br/>
        </para>
        """
        story.append(Paragraph(footer_info, styles['Normal']))
        
        # 签核区
        signature_data = [
            ['制表人:', '_' * 20, '(日期: _' + '_' * 10 + ')'],
            ['审核人:', '_' * 20, '(日期: _' + '_' * 10 + ')'],
            ['批准人:', '_' * 20, '(日期: _' + '_' * 10 + ')'],
        ]
        
        signature_table = Table(signature_data, colWidths=[3*cm, 5*cm, 4*cm])
        signature_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(signature_table)
        
        return story
        
    def _write_excel_info_sheet(self, ws, report_data: ReportData):
        """写入Excel基本信息工作表"""
        # 标题
        ws['A1'] = f"{report_data.workpiece_model}型号管板检测报告"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # 基本信息
        info_data = [
            ('报告ID', report_data.report_id),
            ('产品型号', report_data.workpiece_model),
            ('工件序列号', report_data.workpiece_serial),
            ('检测设备', report_data.detection_equipment),
            ('开始时间', report_data.start_time.strftime('%Y-%m-%d %H:%M:%S') if report_data.start_time else ''),
            ('结束时间', report_data.end_time.strftime('%Y-%m-%d %H:%M:%S') if report_data.end_time else ''),
            ('操作员', report_data.operator),
            ('最终结论', report_data.overall_result),
        ]
        
        for i, (key, value) in enumerate(info_data, 3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            
    def _write_excel_data_sheet(self, ws, report_data: ReportData):
        """写入Excel检测数据工作表"""
        # 表头
        headers = ['孔位ID', '最小直径(mm)', '最大直径(mm)', '平均直径(mm)', '椭圆度', '表面缺陷', '最终状态', '检测时间']
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
            ws.cell(row=1, column=i).font = Font(bold=True)
            
        # 数据
        for row, data in enumerate(report_data.all_data, 2):
            ws.cell(row=row, column=1, value=data['hole_id'])
            ws.cell(row=row, column=2, value=data['min_diameter'])
            ws.cell(row=row, column=3, value=data['max_diameter'])
            ws.cell(row=row, column=4, value=data['avg_diameter'])
            ws.cell(row=row, column=5, value=data['ellipticity'])
            ws.cell(row=row, column=6, value='有' if data['has_defect'] else '无')
            ws.cell(row=row, column=7, value='合格' if data['qualified'] else '不合格')
            ws.cell(row=row, column=8, value=data['detection_time'])
            
    def _write_excel_stats_sheet(self, ws, report_data: ReportData):
        """写入Excel统计分析工作表"""
        # 统计摘要
        stats_data = [
            ('统计项目', '数值'),
            ('总孔数', report_data.total_holes),
            ('已检孔数', report_data.checked_holes),
            ('合格孔数', report_data.qualified_holes),
            ('不合格孔数', report_data.unqualified_holes),
            ('合格率(%)', round(report_data.qualification_rate, 2)),
        ]
        
        for i, (key, value) in enumerate(stats_data, 1):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
            if i == 1:  # 表头
                ws.cell(row=i, column=1).font = Font(bold=True)
                ws.cell(row=i, column=2).font = Font(bold=True)
                
    def _write_excel_nc_sheet(self, ws, report_data: ReportData):
        """写入Excel不合格项工作表"""
        # 表头
        headers = ['序号', '孔位ID', '问题类型', '机检结果', '复检人', '复检结果', '最终判定']
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
            ws.cell(row=1, column=i).font = Font(bold=True)
            
        # 不合格项数据
        for row, nc in enumerate(report_data.non_conformities, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=nc['hole_id'])
            ws.cell(row=row, column=3, value=nc['problem_type'])
            ws.cell(row=row, column=4, value=nc['measurement_result'])
            
            review = nc.get('review_record', {})
            ws.cell(row=row, column=5, value=review.get('reviewer', ''))
            ws.cell(row=row, column=6, value=review.get('review_result', ''))
            ws.cell(row=row, column=7, value=review.get('final_decision', ''))


# 报告生成线程（用于避免阻塞UI）
class ReportGenerationThread(QThread):
    """报告生成线程"""
    
    progress_updated = Signal(int)
    status_updated = Signal(str)
    generation_completed = Signal(str, str)  # 报告类型, 文件路径
    generation_failed = Signal(str, str)     # 报告类型, 错误信息
    
    def __init__(self, generator: ReportGenerator, report_type: str, 
                 hole_data: Dict, workpiece_info: Dict):
        super().__init__()
        self.generator = generator
        self.report_type = report_type
        self.hole_data = hole_data
        self.workpiece_info = workpiece_info
        
    def run(self):
        """执行报告生成"""
        try:
            self.status_updated.emit(f"开始生成{self.report_type}报告...")
            self.progress_updated.emit(10)
            
            if self.report_type == "PDF":
                self.progress_updated.emit(30)
                file_path = self.generator.generate_static_pdf_report(
                    self.hole_data, self.workpiece_info
                )
                self.progress_updated.emit(100)
                self.generation_completed.emit("PDF", file_path)
                
            elif self.report_type == "Excel":
                self.progress_updated.emit(30)
                file_path = self.generator.export_raw_data_excel(
                    self.hole_data, self.workpiece_info
                )
                self.progress_updated.emit(100)
                self.generation_completed.emit("Excel", file_path)
                
            elif self.report_type == "CSV":
                self.progress_updated.emit(30)
                file_path = self.generator.export_raw_data_csv(
                    self.hole_data, self.workpiece_info
                )
                self.progress_updated.emit(100)
                self.generation_completed.emit("CSV", file_path)
                
            else:
                raise ValueError(f"不支持的报告类型: {self.report_type}")
                
        except Exception as e:
            self.generation_failed.emit(self.report_type, str(e))


if __name__ == "__main__":
    """测试报告生成功能"""
    
    # 创建报告生成器
    generator = ReportGenerator()
    
    # 准备测试数据
    workpiece_info = {
        'model': 'CP1400',
        'serial': 'SN-202507-188',
        'operator': '王磊',
        'start_time': datetime.now(),
        'end_time': datetime.now()
    }
    
    hole_data = {
        'total_holes': 1500,
        'measurement_data': []  # 实际应包含测量数据
    }
    
    try:
        # 测试PDF报告生成
        if REPORTLAB_AVAILABLE:
            print("测试PDF报告生成...")
            pdf_path = generator.generate_static_pdf_report(hole_data, workpiece_info)
            print(f"PDF报告已生成: {pdf_path}")
        
        # 测试Excel导出
        if OPENPYXL_AVAILABLE:
            print("测试Excel数据导出...")
            excel_path = generator.export_raw_data_excel(hole_data, workpiece_info)
            print(f"Excel文件已生成: {excel_path}")
        
        # 测试CSV导出
        print("测试CSV数据导出...")
        csv_path = generator.export_raw_data_csv(hole_data, workpiece_info)
        print(f"CSV文件已生成: {csv_path}")
        
        print("✅ 所有测试完成")
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()